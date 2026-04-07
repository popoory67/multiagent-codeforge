"""
This script converts all .docx, .pdf, and .xlsx files in a specified folder to .md format.

Usage:
    python run_convert_to_md.py <input_folder> [-o <output_folder>]

if -o/--output is not specified, the converted .md files will be saved in the same location as the source files.
"""
import sys
import subprocess
import argparse
from pathlib import Path

# Ensure required packages are installed before importing them
#   pip install python-docx pymupdf openpyxl
#   pip install --upgrade python-docx pymupdf openpyxl
def _ensure_packages():
    missing = []
    try:
        import docx  # noqa: F401
    except ImportError:
        missing.append("python-docx")
    try:
        import pymupdf  # noqa: F401
    except ImportError:
        missing.append("pymupdf")
    try:
        import openpyxl  # noqa: F401
    except ImportError:
        missing.append("openpyxl")
    if missing:
        print(f"Installing missing packages: {', '.join(missing)}")
        subprocess.check_call([sys.executable, "-m", "pip", "install", *missing])


_ensure_packages()

from utils.docx_reader import docx_to_md  # noqa: E402
import pymupdf  # noqa: E402
from openpyxl import load_workbook  # noqa: E402


def _table_to_md(table: list[list]) -> str:
    """Convert a 2D list (table) to markdown table."""
    if not table or not table[0]:
        return ""
    rows = []
    for row in table:
        cells = [str(c).strip().replace("\n", " ").replace("|", "\\|") for c in row]
        rows.append("| " + " | ".join(cells) + " |")
    if not rows:
        return ""
    sep = "| " + " | ".join(["---"] * len(table[0])) + " |"
    return rows[0] + "\n" + sep + "\n" + "\n".join(rows[1:])


def xlsx_to_md(xlsx_path: Path) -> str:
    wb = load_workbook(str(xlsx_path), read_only=True, data_only=True)
    sections = []
    for sheet in wb.worksheets:
        rows = []
        for row in sheet.iter_rows(values_only=True):
            cells = [str(c).strip().replace("\n", " ").replace("|", "\\|") if c is not None else "" for c in row]
            rows.append(cells)
        if not rows:
            continue
        # Drop fully empty trailing rows
        while rows and all(c == "" for c in rows[-1]):
            rows.pop()
        if not rows:
            continue
        col_count = max(len(r) for r in rows)
        lines = [f"## {sheet.title}", ""]
        for i, row in enumerate(rows):
            # Pad rows to consistent column count
            padded = row + [""] * (col_count - len(row))
            lines.append("| " + " | ".join(padded) + " |")
            if i == 0:
                lines.append("| " + " | ".join(["---"] * col_count) + " |")
        sections.append("\n".join(lines))
    wb.close()
    return "\n\n".join(sections)


def pdf_to_md(pdf_path: Path, images_dir: Path | None = None, img_rel_path: str = "") -> str:
    doc = pymupdf.open(str(pdf_path))
    pages = []
    img_counter = 1
    seen_xrefs = set()

    for page in doc:
        parts = []

        # Extract tables first
        tables = page.find_tables()
        table_rects = []
        for table in tables:
            table_rects.append(table.bbox)
            data = table.extract()
            if data:
                parts.append(("table", table.bbox[1], _table_to_md(data)))

        # Extract text blocks and images
        blocks = page.get_text("blocks")  # (x0, y0, x1, y1, text, block_no, block_type)
        for b in blocks:
            by0, by1 = b[1], b[3]

            if b[6] == 1 and images_dir:
                # Image block — extract via page.get_images
                continue  # handled below
            elif b[6] != 0:
                continue

            # Check if this block overlaps with any table
            in_table = False
            for tr in table_rects:
                if by0 >= tr[1] - 2 and by1 <= tr[3] + 2:
                    in_table = True
                    break
            if not in_table:
                text = b[4].strip()
                if text:
                    parts.append(("text", by0, text))

        # Extract images from page
        if images_dir:
            for img_info in page.get_images(full=True):
                xref = img_info[0]
                if xref in seen_xrefs:
                    continue
                seen_xrefs.add(xref)
                extracted = doc.extract_image(xref)
                if not extracted or not extracted.get("image"):
                    continue
                ext = extracted.get("ext", "png")
                img_name = f"image_{img_counter:03d}.{ext}"
                img_counter += 1
                (images_dir / img_name).write_bytes(extracted["image"])
                # Place image at approximate vertical position (use page midpoint as fallback)
                parts.append(("image", 0, f"![{img_name}]({img_rel_path}/{img_name})"))

        # Sort by vertical position to maintain reading order
        parts.sort(key=lambda p: p[1])
        page_content = "\n\n".join(p[2] for p in parts)
        if page_content.strip():
            pages.append(page_content)

    doc.close()
    return "\n\n---\n\n".join(pages)


_img_folder_counter = 0


def convert_file(src: Path, out_dir: Path | None) -> Path:
    global _img_folder_counter
    _img_folder_counter += 1
    ext = src.suffix.lower()

    if out_dir:
        out_dir.mkdir(parents=True, exist_ok=True)
        dest = out_dir / src.with_suffix(".md").name
    else:
        dest = src.with_suffix(".md")

    # Create images directory next to the output md file
    img_id = f"img_{_img_folder_counter:03d}"
    images_dir = dest.parent / "images" / img_id
    images_dir.mkdir(parents=True, exist_ok=True)

    if ext == ".docx":
        md = docx_to_md(src, images_dir=images_dir, img_rel_path=f"images/{img_id}")
    elif ext == ".pdf":
        md = pdf_to_md(src, images_dir=images_dir, img_rel_path=f"images/{img_id}")
    elif ext in (".xlsx", ".xlsm"):
        md = xlsx_to_md(src)
    else:
        raise ValueError(f"Unsupported format: {ext}")

    # Remove images dir if no images were extracted
    if not any(images_dir.iterdir()):
        images_dir.rmdir()

    dest.write_text(md, encoding="utf-8")
    return dest


def main():
    sys.stdout.reconfigure(encoding="utf-8")

    parser = argparse.ArgumentParser(description="Convert docx/pdf/xlsx files in a folder to md format")
    parser.add_argument("input_folder", help="Path to the folder containing files to convert")
    parser.add_argument("-o", "--output", help="Output folder (if not specified, files will be saved in the same location)")
    parser.add_argument("-b", "--blacklist", nargs="*", default=[], help="Directory names to exclude from conversion")
    args = parser.parse_args()

    src_dir = Path(args.input_folder)
    if not src_dir.is_dir():
        print(f"Folder not found: {src_dir}")
        sys.exit(1)

    out_dir = Path(args.output) if args.output else None
    blacklist = set(args.blacklist)

    files = sorted(
        f for f in src_dir.rglob("*")
        if f.suffix.lower() in (".docx", ".pdf", ".xlsx", ".xlsm")
        and not f.name.startswith("~")
        and not any(bl in part for bl in blacklist for part in f.relative_to(src_dir).parts)
        and not any(bl in f.name for bl in blacklist)
    )

    if not files:
        print("No docx/pdf/xlsx files to convert.")
        sys.exit(0)

    print(f"Total {len(files)} files conversion started\n")

    for f in files:
        try:
            dest = convert_file(f, out_dir)
            print(f"  ✓ {f.name} → {dest}")
        except Exception as e:
            print(f"  ✗ {f.name} — {e}")

    print("\nConversion completed.")


if __name__ == "__main__":
    main()

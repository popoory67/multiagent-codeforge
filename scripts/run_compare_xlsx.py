"""
Compare two Excel files by a key column to detect changes, additions, and deletions.

Usage:
    python run_compare_xlsx.py old.xlsx new.xlsx -s "Sheet" --start 8 J K M --key J
    python run_compare_xlsx.py old.xlsx new.xlsx -s "Sheet" --start 8 M --key J --info K L

The --start option works the same as run_convert_to_md.py:
    First arg = header row, rest = columns to compare.
The --key option specifies which column is used as the row identifier (name).
The --info option specifies additional identifier columns shown on the left (not compared).
"""
import sys
import argparse
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string


def _parse_col(col_str: str) -> int:
    """Convert column letter or number string to 1-based index."""
    return int(col_str) if col_str.isdigit() else column_index_from_string(col_str.upper())


def read_xlsx_data(xlsx_path: Path, sheet_name: str, header_row: int, col_indices: list[int]) -> tuple[list[str], list[list[str]]]:
    """Read xlsx and return (headers, rows) for the given columns.

    Returns:
        headers: list of header strings
        rows: list of rows, each row is a list of cell strings
    """
    wb = load_workbook(str(xlsx_path), read_only=True, data_only=True)
    sheets = [s for s in wb.worksheets if s.title == sheet_name]
    if not sheets:
        available = [s.title for s in wb.worksheets]
        wb.close()
        raise ValueError(f"Sheet '{sheet_name}' not found. Available: {available}")

    all_rows = list(sheets[0].iter_rows(values_only=True))
    wb.close()

    if header_row - 1 >= len(all_rows):
        raise ValueError(f"Header row {header_row} exceeds sheet row count {len(all_rows)}")

    zero_cols = [c - 1 for c in col_indices]

    def pick(row_data):
        return [
            str(row_data[i]).strip() if i < len(row_data) and row_data[i] is not None else ""
            for i in zero_cols
        ]

    headers = pick(all_rows[header_row - 1])
    rows = []
    for row_data in all_rows[header_row:]:
        rows.append(pick(row_data))

    # Drop trailing empty rows
    while rows and all(c == "" for c in rows[-1]):
        rows.pop()

    return headers, rows


def _esc(val: str) -> str:
    """Escape characters that break markdown tables."""
    return val.replace("|", "\\|").replace("\n", "<br>")


def compare(old_path: Path, new_path: Path, sheet_name: str, header_row: int,
            col_indices: list[int], key_col: int, info_cols: list[int] | None = None) -> str:
    """Compare two xlsx files and return a markdown diff report.

    Args:
        col_indices: columns to compare (NOT including key_col or info_cols)
        key_col: the primary key column for matching rows
        info_cols: additional identifier columns shown on the left (not compared)
    """
    info_cols = info_cols or []
    # Read: key + info + compare columns
    left_cols = [key_col] + info_cols
    all_cols = left_cols + col_indices
    old_headers, old_rows = read_xlsx_data(old_path, sheet_name, header_row, all_cols)
    _, new_rows = read_xlsx_data(new_path, sheet_name, header_row, all_cols)

    n_left = len(left_cols)  # key + info count
    left_headers = old_headers[:n_left]
    compare_headers = old_headers[n_left:]

    # Build dicts: key -> (row_num, left_vals, compare_vals)
    def build_dict(rows, start_row):
        d = {}
        for idx, row in enumerate(rows):
            name = row[0]
            if name:
                d[name] = (start_row + idx, row[:n_left], row[n_left:])
        return d

    old_dict = build_dict(old_rows, header_row + 1)
    new_dict = build_dict(new_rows, header_row + 1)

    old_names = set(old_dict.keys())
    new_names = set(new_dict.keys())

    added = sorted(new_names - old_names)
    deleted = sorted(old_names - new_names)
    common = sorted(old_names & new_names)

    # Build table rows: (old_row, new_row, left_vals, status, column, old_val, new_val)
    table_rows = []
    all_names = sorted(old_names | new_names)

    for name in all_names:
        if name in common:
            old_rn, _, old_cmp = old_dict[name]
            new_rn, new_left, new_cmp = new_dict[name]
            left_vals = new_left
            has_change = False
            for i, header in enumerate(compare_headers):
                old_val = old_cmp[i] if i < len(old_cmp) else ""
                new_val = new_cmp[i] if i < len(new_cmp) else ""
                if old_val != new_val:
                    has_change = True
                    table_rows.append((old_rn, new_rn, left_vals, "Changed", header, old_val, new_val))
            if not has_change:
                pass  # unchanged rows are omitted
        elif name in new_names:
            new_rn, left_vals, cmp_vals = new_dict[name]
            for i, header in enumerate(compare_headers):
                val = cmp_vals[i] if i < len(cmp_vals) else ""
                table_rows.append(("", new_rn, left_vals, "Added", header, "", val))
        else:
            old_rn, left_vals, cmp_vals = old_dict[name]
            for i, header in enumerate(compare_headers):
                val = cmp_vals[i] if i < len(cmp_vals) else ""
                table_rows.append((old_rn, "", left_vals, "Deleted", header, val, ""))

    # Build markdown report
    lines = ["# Comparison Report", ""]
    lines.append(f"- **Old file**: `{old_path.name}`")
    lines.append(f"- **New file**: `{new_path.name}`")
    lines.append(f"- **Sheet**: {sheet_name}")
    lines.append("")

    n_changed = len(set(lv[0] for _, _, lv, s, *_ in table_rows if s == "Changed"))
    n_added = len(added)
    n_deleted = len(deleted)
    lines.append(f"Changed: {n_changed} / Added: {n_added} / Deleted: {n_deleted} / Unchanged: {len(common) - n_changed}")
    lines.append("")

    # Table of contents
    lines.append("## Table of Contents")
    lines.append("")
    if any(r[3] == "Changed" for r in table_rows):
        lines.append(f"- [Changed ({n_changed})](#changed)")
    if any(r[3] == "Added" for r in table_rows):
        lines.append(f"- [Added ({n_added})](#added)")
    if any(r[3] == "Deleted" for r in table_rows):
        lines.append(f"- [Deleted ({n_deleted})](#deleted)")
    lines.append("")

    if table_rows:
        left_hdr = " | ".join(_esc(h) for h in left_headers)
        header_line = f"| Old Row | New Row | {left_hdr} | Status | Column | Old | New |"
        sep_line = "| ---: | ---: | " + " | ".join(["---"] * n_left) + " | --- | --- | --- | --- |"

        for section_status, section_title in [("Changed", "Changed"), ("Added", "Added"), ("Deleted", "Deleted")]:
            section_rows = [r for r in table_rows if r[3] == section_status]
            if not section_rows:
                continue
            lines.append(f"## {section_title}")
            lines.append("")
            lines.append(header_line)
            lines.append(sep_line)
            for old_rn, new_rn, left_vals, status, col, old_val, new_val in section_rows:
                left_str = " | ".join(_esc(v) for v in left_vals)
                lines.append(f"| {old_rn} | {new_rn} | {left_str} | {status} | {_esc(col)} | {_esc(old_val)} | {_esc(new_val)} |")
            lines.append("")
    else:
        lines.append("No differences found.")

    lines.append("")
    return "\n".join(lines)


def main():
    sys.stdout.reconfigure(encoding="utf-8")

    parser = argparse.ArgumentParser(description="Compare two Excel files by key column")
    parser.add_argument("old_file", help="Path to the old (base) xlsx file")
    parser.add_argument("new_file", help="Path to the new (updated) xlsx file")
    parser.add_argument("-s", "--sheet", required=True, help="Sheet (tab) name")
    parser.add_argument("--start", nargs="+", required=True, metavar="ARG",
                        help="Header row and columns, e.g. --start 8 J K M")
    parser.add_argument("--key", required=True,
                        help="Key column (name column) for matching rows, e.g. J")
    parser.add_argument("--info", nargs="*", default=[], metavar="COL",
                        help="Info columns shown on the left (not compared), e.g. --info K L")
    parser.add_argument("-o", "--output", default=None, help="Output markdown file path")
    args = parser.parse_args()

    old_path = Path(args.old_file)
    new_path = Path(args.new_file)
    if not old_path.exists():
        print(f"File not found: {old_path}")
        sys.exit(1)
    if not new_path.exists():
        print(f"File not found: {new_path}")
        sys.exit(1)

    header_row = int(args.start[0])
    key_col = _parse_col(args.key)
    info_cols = [_parse_col(c) for c in args.info]
    skip = {key_col} | set(info_cols)
    col_indices = [_parse_col(c) for c in args.start[1:] if _parse_col(c) not in skip]

    report = compare(old_path, new_path, args.sheet, header_row, col_indices, key_col, info_cols)

    if args.output:
        out = Path(args.output)
        out.parent.mkdir(parents=True, exist_ok=True)
        out.write_text(report, encoding="utf-8")
        print(f"Report saved to {out}")
    else:
        print(report)


if __name__ == "__main__":
    main()
